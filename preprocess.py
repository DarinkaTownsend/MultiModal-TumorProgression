import os
import sys
import csv
import glob
import numpy as np
import openpyxl
from collections import defaultdict

# ============================================================
# UCSF-ALPTDG Preprocessing for Tumor Severity Classification
# Extracts 2D axial slices from NIfTI volumes, pairs with
# clinical metadata, builds text prompts, outputs manifest CSV
# ============================================================

# paths — adjust these to match your Newton layout
DATA_DIR = os.path.expanduser("~/tumor_progression/data")
XLSX_PATH = os.path.join(DATA_DIR, "UCSF_PostopGlioma_Table S1 R1 V5.0_UNBLINDED_FINAL.xlsx")
OUTPUT_DIR = os.path.expanduser("~/tumor_progression/processed")
SLICE_DIR = os.path.join(OUTPUT_DIR, "slices")

# grade grouping: 3-class severity
# low = grades 1,2 | mid = grade 3 | high = grade 4
GRADE_TO_SEVERITY = {1: 0, 2: 0, 3: 1, 4: 2}
SEVERITY_NAMES = {0: "low", 1: "mid", 2: "high"}

# MRI modalities to extract per timepoint
MODALITIES = ["flair", "t1", "t1ce", "t2"]

# only keep axial slices where tumor covers at least this
# fraction of nonzero voxels in the slice
MIN_TUMOR_PIXELS = 10


def loadClinicalInfo(xlsxPath):
    # returns dict: subjectId -> {diagnosis, grade, sex, age, idh, mgmt, ...}
    wb = openpyxl.load_workbook(xlsxPath)
    ws = wb["Clinical Info"]
    headers = [c.value for c in ws[1]]
    clinical = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        subjectId = row[0]
        if subjectId is None:
            continue
        clinical[int(subjectId)] = {
            "subjectId": int(subjectId),
            "daysSurgeryToScan": row[1],
            "daysBetweenScans": row[2],
            "extentOfResection": row[3],
            "diagnosis": row[4],
            "presumedDiagnosis": row[5],
            "grade": row[6],
            "mgmt": row[7],
            "mgmtIndex": row[8],
            "idh": row[9],
            "oneP19q": row[10],
            "atrx": row[11],
            "daysDeathToScan": row[12],
            "daysToProgression": row[13],
            "daysToRT": row[14],
            "chemoType": row[15],
            "daysChemoToScan": row[16],
            "txAtScan1": row[17],
            "txAtScan2": row[18],
            "numSurgeries": row[19],
        }
    return clinical


def loadTrainTestSplit(xlsxPath):
    # returns dict: subjectId -> "TRAIN" or "TEST"
    wb = openpyxl.load_workbook(xlsxPath)
    ws = wb["TrainTestSplit"]
    splits = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            splits[int(row[0])] = row[1]
    return splits


def loadImagingMeta(xlsxPath):
    # returns dict: (subjectId, timepoint) -> {sex, age, scanner, volumes...}
    wb = openpyxl.load_workbook(xlsxPath)
    ws = wb.active  # first sheet has per-timepoint imaging data
    imaging = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        subjectId = row[0]
        timepoint = row[1]
        if subjectId is None:
            continue
        imaging[(int(subjectId), int(timepoint))] = {
            "sex": row[6],
            "age": row[7],
            "ncrVolume": row[8],
            "snfhVolume": row[9],
            "etVolume": row[10],
            "rcVolume": row[11],
            "wtVolume": row[12],
            "tcVolume": row[13],
            "scanner": row[5],
        }
    return imaging


def buildClinicalText(clinInfo, imagingInfo):
    # constructs a natural language clinical summary for the text encoder
    parts = []

    # demographics
    age = imagingInfo.get("age", None) if imagingInfo else None
    sex = imagingInfo.get("sex", None) if imagingInfo else None
    if age and sex:
        parts.append("%d year old %s" % (int(age), str(sex).lower()))

    # diagnosis
    diag = clinInfo.get("diagnosis", None)
    grade = clinInfo.get("grade", None)
    if diag:
        if grade:
            parts.append("WHO grade %d %s" % (int(grade), diag))
        else:
            parts.append(diag)

    # molecular markers
    idh = clinInfo.get("idh", None)
    if idh:
        parts.append("IDH %s" % idh)

    mgmt = clinInfo.get("mgmt", None)
    if mgmt:
        parts.append("MGMT %s" % mgmt)

    oneP = clinInfo.get("oneP19q", None)
    if oneP:
        parts.append("1p19q %s" % oneP)

    atrx = clinInfo.get("atrx", None)
    if atrx:
        parts.append("ATRX %s" % atrx)

    # treatment
    eor = clinInfo.get("extentOfResection", None)
    if eor:
        parts.append("extent of resection: %s" % eor)

    chemo = clinInfo.get("chemoType", None)
    if chemo:
        parts.append("chemotherapy: %s" % chemo)

    tx1 = clinInfo.get("txAtScan1", None)
    if tx1:
        parts.append("treatment at scan 1: %s" % tx1)

    # time context
    daysBetween = clinInfo.get("daysBetweenScans", None)
    if daysBetween:
        parts.append("interval between scans: %d days" % int(daysBetween))

    return ", ".join(parts) if parts else "no clinical information available"


def getSeverityLabel(grade):
    # maps grade to 3-class severity, returns None if unmappable
    if grade is None:
        return None
    g = int(grade)
    return GRADE_TO_SEVERITY.get(g, None)


def extractSlices(niftiPath):
    # loads a nifti volume and returns all axial slices as numpy arrays
    import nibabel as nib
    img = nib.load(niftiPath)
    vol = img.get_fdata()
    slices = []
    for i in range(vol.shape[2]):
        slices.append(vol[:, :, i].astype(np.float32))
    return slices


def findTumorSlices(segPath):
    # returns set of axial slice indices that contain tumor
    import nibabel as nib
    seg = nib.load(segPath)
    vol = seg.get_fdata()
    tumorSlices = set()
    for i in range(vol.shape[2]):
        sliceData = vol[:, :, i]
        if np.sum(sliceData > 0) >= MIN_TUMOR_PIXELS:
            tumorSlices.add(i)
    return tumorSlices, vol.shape[2]


def processPatient(subjectId, dataDir, outputDir):
    # extracts 2D slices for all modalities at both timepoints
    # returns list of (sliceIdx, time1Paths, time2Paths, segPaths)
    patientDir = os.path.join(dataDir, str(subjectId))
    if not os.path.isdir(patientDir):
        return []

    prefix = str(subjectId)

    # check which files exist
    time1SegPath = os.path.join(patientDir, "%s_time1_seg.nii.gz" % prefix)
    time2SegPath = os.path.join(patientDir, "%s_time2_seg.nii.gz" % prefix)

    if not os.path.exists(time1SegPath) or not os.path.exists(time2SegPath):
        print("  [SKIP] missing seg for %s" % prefix)
        return []

    # find slices with tumor in either timepoint
    tumorSlices1, nSlices1 = findTumorSlices(time1SegPath)
    tumorSlices2, nSlices2 = findTumorSlices(time2SegPath)

    if nSlices1 != nSlices2:
        print("  [WARN] slice count mismatch for %s: %d vs %d" % (prefix, nSlices1, nSlices2))

    # union of tumor slices from both timepoints
    relevantSlices = sorted(tumorSlices1 | tumorSlices2)

    if len(relevantSlices) == 0:
        print("  [SKIP] no tumor slices for %s" % prefix)
        return []

    # create output directory for this patient
    patientOutDir = os.path.join(outputDir, prefix)
    os.makedirs(patientOutDir, exist_ok=True)

    records = []

    # load all volumes once
    import nibabel as nib

    volumes = {}
    for tp in [1, 2]:
        tpStr = "time%d" % tp
        for mod in MODALITIES:
            # handle the t1ce naming: file is named t1ce not t1ce
            fpath = os.path.join(patientDir, "%s_%s_%s.nii.gz" % (prefix, tpStr, mod))
            if not os.path.exists(fpath):
                print("  [WARN] missing %s" % fpath)
                volumes[(tp, mod)] = None
                continue
            volumes[(tp, mod)] = nib.load(fpath).get_fdata()

        # load seg
        segPath = os.path.join(patientDir, "%s_%s_seg.nii.gz" % (prefix, tpStr))
        volumes[(tp, "seg")] = nib.load(segPath).get_fdata()

    # extract and save relevant slices
    for sliceIdx in relevantSlices:
        slicePaths = {}
        valid = True

        for tp in [1, 2]:
            tpStr = "time%d" % tp
            for mod in MODALITIES + ["seg"]:
                vol = volumes.get((tp, mod), None)
                if vol is None:
                    valid = False
                    break
                if sliceIdx >= vol.shape[2]:
                    valid = False
                    break

                sliceData = vol[:, :, sliceIdx].astype(np.float32)
                fname = "%s_%s_%s_s%03d.npy" % (prefix, tpStr, mod, sliceIdx)
                outPath = os.path.join(patientOutDir, fname)
                np.save(outPath, sliceData)
                slicePaths["%s_%s" % (tpStr, mod)] = outPath

            if not valid:
                break

        if valid:
            records.append({
                "subjectId": subjectId,
                "sliceIdx": sliceIdx,
                "totalSlices": nSlices1,
                "paths": slicePaths,
            })

    return records


def main():
    print("Loading clinical data...")
    clinical = loadClinicalInfo(XLSX_PATH)
    print("  %d patients with clinical info" % len(clinical))

    print("Loading train/test split...")
    splits = loadTrainTestSplit(XLSX_PATH)
    print("  %d patients in split table" % len(splits))

    print("Loading imaging metadata...")
    imaging = loadImagingMeta(XLSX_PATH)
    print("  %d timepoint records" % len(imaging))

    # filter to patients with valid severity labels
    validPatients = {}
    for sid, info in clinical.items():
        severity = getSeverityLabel(info.get("grade", None))
        if severity is not None:
            validPatients[sid] = severity
    print("  %d patients with valid grade labels" % len(validPatients))

    # create output dirs
    os.makedirs(SLICE_DIR, exist_ok=True)

    # process each patient
    allRecords = []
    severityCounts = defaultdict(int)

    for idx, (sid, severity) in enumerate(sorted(validPatients.items())):
        print("[%d/%d] Processing patient %d..." % (idx + 1, len(validPatients), sid))

        records = processPatient(sid, DATA_DIR, SLICE_DIR)

        clinInfo = clinical.get(sid, {})
        imagingInfo = imaging.get((sid, 1), {})
        clinText = buildClinicalText(clinInfo, imagingInfo)
        split = splits.get(sid, "UNKNOWN")

        for rec in records:
            rec["severity"] = severity
            rec["severityName"] = SEVERITY_NAMES[severity]
            rec["clinicalText"] = clinText
            rec["split"] = split
            rec["grade"] = clinInfo.get("grade", None)
            rec["diagnosis"] = clinInfo.get("diagnosis", None)
            allRecords.append(rec)

        severityCounts[severity] += len(records)

    # write manifest CSV
    manifestPath = os.path.join(OUTPUT_DIR, "manifest.csv")
    print("\nWriting manifest to %s..." % manifestPath)

    fieldnames = [
        "subjectId", "sliceIdx", "totalSlices", "split",
        "grade", "severity", "severityName", "diagnosis",
        "clinicalText",
        "time1_flair", "time1_t1", "time1_t1ce", "time1_t2", "time1_seg",
        "time2_flair", "time2_t1", "time2_t1ce", "time2_t2", "time2_seg",
    ]

    with open(manifestPath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for rec in allRecords:
            row = {
                "subjectId": rec["subjectId"],
                "sliceIdx": rec["sliceIdx"],
                "totalSlices": rec["totalSlices"],
                "split": rec["split"],
                "grade": rec["grade"],
                "severity": rec["severity"],
                "severityName": rec["severityName"],
                "diagnosis": rec["diagnosis"],
                "clinicalText": rec["clinicalText"],
            }
            for key, path in rec["paths"].items():
                row[key] = path
            writer.writerow(row)

    # summary
    print("\n=== Summary ===")
    print("Total slices: %d" % len(allRecords))
    for sev in sorted(severityCounts.keys()):
        print("  %s (class %d): %d slices" % (SEVERITY_NAMES[sev], sev, severityCounts[sev]))

    trainCount = sum(1 for r in allRecords if r["split"] == "TRAIN")
    testCount = sum(1 for r in allRecords if r["split"] == "TEST")
    print("  Train slices: %d" % trainCount)
    print("  Test slices: %d" % testCount)
    print("\nDone. Manifest: %s" % manifestPath)
    print("Slices saved to: %s" % SLICE_DIR)


if __name__ == "__main__":
    main()

